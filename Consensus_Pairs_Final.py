#!/usr/bin/env python
# coding: utf-8


from sklearn.cluster import kmeans_plusplus
from sklearn.datasets import make_blobs
import matplotlib.pyplot as plt
import math
import xlsxwriter
import pandas as pd
from openpyxl import load_workbook


wb = load_workbook(filename='AHPcalc-Decisions.xlsx', 
                   read_only=True, data_only=True)

numParticipants = wb["Summary"]["B7"].value
numCriteria = wb["Summary"]["B5"].value

decision_matrix = [] # matrix including the priority percentages for all criteria and participants

for i in range(numParticipants):
    decision_matrix_list = [] # list containing the judgements of each participant
    my_range = wb.defined_names['RGGM{}'.format(i + 1)].value # we get a string representing the array with
                                                              # the resulting priorities for the i-th participant
    for j in range(numCriteria): # loop to retrieve each cell from the previous array and add the value
                                 # and add it to decision_matrix_list
        if '!' in my_range:
                # passed a worksheet!cell reference
                ws_name, reg = my_range.split('!')
                if ws_name.startswith("'") and ws_name.endswith("'"):
                    # optionally strip single quotes around sheet name
                    ws_name = ws_name[1:-1]
                region = wb[ws_name][reg]
        decision_matrix_list.append(region[j][0].value)
    decision_matrix.append(decision_matrix_list)


pair_participants = 2
num_comparisons = numParticipants * (numParticipants - 1) / 2
workbook = xlsxwriter.Workbook('Consensus_Decisions.xlsx')
worksheet = workbook.add_worksheet()

for n in range(numParticipants):
    for m in range(numParticipants):
        worksheet.write(n + 1, 0, "Participant {}".format(n + 1))
        worksheet.write(0, m + 1, "Participant {}".format(m + 1))
        if(n < m):
            m_pair = [[decision_matrix[i][j] for j in range(numCriteria)] for i in range(numParticipants) if i == n or i == m] # matrix containing the judgements from
                                                                                                                               # a pair of participants
            m_p_avg = [] # list containing the average of priority percentage from both participants for each criterion
            for j in range(numCriteria):
                sum_p_avg = 0
                for i in range(pair_participants):
                    sum_p_avg = sum_p_avg + m_pair[i][j]
                m_p_avg.append(sum_p_avg/pair_participants)
            
            m_entropy = [[-m_pair[i][j]*math.log(m_pair[i][j]) for j in range(numCriteria)] for i in range(pair_participants)] #Shannon entropy matrix for m_pair
            
            ln_m_p_avg = [-m_p_avg[i]*math.log(m_p_avg[i]) for i in range(len(m_p_avg))] # list for the gamma entropy
            
            m_sum_alpha = [] # list for the alpha entropy

            for i in range(pair_participants):
                m_sum_alpha.append(sum(m_entropy[i]))
            
            h_gamma = math.exp(sum(ln_m_p_avg))
            h_alpha = math.exp(sum(m_sum_alpha)/pair_participants)
            h_beta = h_gamma / h_alpha

            h_min_max = numCriteria / math.exp((-9 / (numCriteria + 8) * math.log(9 / (numCriteria + 8)) - 
                                                (numCriteria - 1) * (1 / (numCriteria + 8) * 
                                                                     math.log(1 / (numCriteria + 8))))) # this variable includes the division between
                                                                                                        # the exponential of the maximum gamma entropy
                                                                                                        # (equivalent as the number of criteria) and
                                                                                                        # the exponential of the minimum alpha entropy

            consensus = (1 / h_beta - 1 / h_min_max) / (1 - 1 / h_min_max)
            
            worksheet.write(n + 1, m + 1, consensus)
            worksheet.write(m + 1, n + 1, consensus)

            print("The consensus between ",n," and ",m," is: ", consensus)
            
        worksheet.write(n + 1, n + 1, 1)
        
worksheet.conditional_format(1, 1, numParticipants + 1, numParticipants + 1, {'type': '2_color_scale'})

workbook.close()


pair_criteria = 2
num_comparisons = numCriteria * (numCriteria - 1) / 2
workbook = xlsxwriter.Workbook('Consensus_Decisions.xlsx')
worksheet = workbook.add_worksheet()

for n in range(numCriteria):
    for m in range(numCriteria):
        worksheet.write(n + 1, 0, "Criteria {}".format(n + 1))
        worksheet.write(0, m + 1, "Criteria {}".format(m + 1))
        if(n < m):
            l_pair = [[l[i][j] for j in range(numCriteria) if j == n or j == m] for i in range(numParticipants)]
            l_p_avg = []
            for j in range(pair_criteria):
                sum_p_avg = 0
                for i in range(numParticipants):
                    sum_p_avg = sum_p_avg + l_pair[i][j]
                l_p_avg.append(sum_p_avg/numParticipants)
    
            m_entropy = [[-l_pair[i][j]*math.log(l_pair[i][j]) for j in range(pair_criteria)] for i in range(numParticipants)]
        
            ln_l_p_avg = [-l_p_avg[i]*math.log(l_p_avg[i]) for i in range(len(l_p_avg))]
            
            l_sum_alpha = []

            for i in range(pair_criteria):
                l_sum_alpha.append(sum(m_entropy[i]))
            
            h_gamma = math.exp(sum(ln_l_p_avg))
            
            h_alpha = math.exp(sum(l_sum_alpha)/numParticipants)
            
            h_beta = h_gamma / h_alpha

            h_min_max = pair_criteria / math.exp((-9 / (pair_criteria + 8) * math.log(9 / (pair_criteria + 8)) - (pair_criteria - 1) * (1 / (pair_criteria + 8) * math.log(1 / (pair_criteria + 8)))))
            
            consensus = (1 / h_beta - 1 / h_min_max) / (1 - 1 / h_min_max)
            
            worksheet.write(n + 1, m + 1, consensus)
            worksheet.write(m + 1, n + 1, consensus)

            print("The consensus between ",n," and ",m," is: ", consensus)
            
        worksheet.write(n + 1, n + 1, 1)
        
worksheet.conditional_format(1, 1, numParticipants + 1, numParticipants + 1, {'type': '2_color_scale'})

workbook.close()

